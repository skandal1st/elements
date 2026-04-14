"""Роуты /it/equipment — IT-оборудование."""

import io
import logging
from datetime import datetime
from typing import List, Optional
from uuid import UUID

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import or_
from sqlalchemy.orm import Session

from backend.modules.hr.models.user import User
from backend.modules.hr.models.employee import Employee
from backend.modules.it.dependencies import get_current_user, get_db, require_it_roles
from backend.modules.it.models import (
    Brand,
    Building,
    Consumable,
    Equipment,
    EquipmentHistory,
    EquipmentModel,
    EquipmentType,
    LicenseAssignment,
    ModelConsumable,
    Room,
    SoftwareLicense,
)
from backend.modules.it.schemas.equipment import (
    EquipmentCreate,
    EquipmentOut,
    EquipmentSyncFromScan,
    EquipmentUpdate,
    ScanComputerRequest,
)
from backend.modules.it.services.computer_scanner import get_scan_config, run_scan
from backend.modules.it.schemas.equipment_history import ChangeOwnerRequest

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/equipment", tags=["equipment"])


@router.get(
    "/",
    response_model=List[EquipmentOut],
    dependencies=[Depends(require_it_roles(["admin", "it_specialist", "employee", "auditor"]))],
)
def list_equipment(
    db: Session = Depends(get_db),
    status: Optional[str] = Query(None),
    category: Optional[str] = Query(None),
    owner_id: Optional[int] = Query(None),
    room_id: Optional[UUID] = Query(None),
    search: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
) -> List[EquipmentOut]:
    q = db.query(Equipment)
    if status:
        q = q.filter(Equipment.status == status)
    if category:
        q = q.filter(Equipment.category == category)
    if owner_id:
        q = q.filter(Equipment.current_owner_id == owner_id)
    if room_id:
        q = q.filter(Equipment.room_id == room_id)
    if search and search.strip():
        s = f"%{search.strip()}%"
        q = q.filter(
            or_(
                Equipment.name.ilike(s),
                Equipment.inventory_number.ilike(s),
                Equipment.serial_number.ilike(s),
            )
        )
    q = q.order_by(Equipment.created_at.desc())
    offset = (page - 1) * page_size
    equipment_list = q.offset(offset).limit(page_size).all()
    
    # Собираем все room_id для одного запроса
    room_ids = [eq.room_id for eq in equipment_list if eq.room_id]
    rooms_map = {}
    if room_ids:
        rooms = db.query(Room).filter(Room.id.in_(room_ids)).all()
        building_ids = [r.building_id for r in rooms if r.building_id]
        buildings_map = {}
        if building_ids:
            buildings = db.query(Building).filter(Building.id.in_(building_ids)).all()
            buildings_map = {b.id: b.name for b in buildings}
        rooms_map = {r.id: (r.name, buildings_map.get(r.building_id)) for r in rooms}

    # Собираем владельцев (employees.id) для одного запроса
    owner_ids = [eq.current_owner_id for eq in equipment_list if eq.current_owner_id]
    owners_map = {}
    if owner_ids:
        owners = db.query(Employee).filter(Employee.id.in_(owner_ids)).all()
        owners_map = {o.id: (o.full_name, o.email) for o in owners}
    
    # Формируем результат с информацией о кабинете
    result = []
    for eq in equipment_list:
        eq_out = EquipmentOut.model_validate(eq)
        if eq.room_id and eq.room_id in rooms_map:
            room_name, building_name = rooms_map[eq.room_id]
            eq_out.room_name = room_name
            eq_out.building_name = building_name
        if eq.current_owner_id and eq.current_owner_id in owners_map:
            owner_name, owner_email = owners_map[eq.current_owner_id]
            eq_out.owner_name = owner_name
            eq_out.owner_email = owner_email
        result.append(eq_out)
    
    return result


@router.get(
    "/employee/{employee_id}",
    response_model=List[EquipmentOut],
    dependencies=[Depends(require_it_roles(["admin", "it_specialist"]))],
)
def list_employee_equipment(
    employee_id: int,
    db: Session = Depends(get_db),
) -> List[dict]:
    """Получить оборудование сотрудника по employee_id"""
    employee = db.query(Employee).filter(Employee.id == employee_id).first()
    if not employee:
        raise HTTPException(status_code=404, detail="Сотрудник не найден")

    equipment_list = (
        db.query(Equipment)
        .filter(Equipment.current_owner_id == employee_id)
        .all()
    )

    # Собираем все room_id для одного запроса
    room_ids = [eq.room_id for eq in equipment_list if eq.room_id]
    rooms_map = {}
    if room_ids:
        rooms = db.query(Room).filter(Room.id.in_(room_ids)).all()
        building_ids = [r.building_id for r in rooms if r.building_id]
        buildings_map = {}
        if building_ids:
            buildings = db.query(Building).filter(Building.id.in_(building_ids)).all()
            buildings_map = {b.id: b.name for b in buildings}
        rooms_map = {r.id: (r.name, buildings_map.get(r.building_id)) for r in rooms}

    # Формируем результат с обогащенными данными
    result = []
    for eq in equipment_list:
        eq_out = EquipmentOut.model_validate(eq).model_dump()

        if eq.room_id and eq.room_id in rooms_map:
            room_name, building_name = rooms_map[eq.room_id]
            eq_out["room_name"] = room_name
            eq_out["building_name"] = building_name

        eq_out["owner_name"] = employee.full_name
        eq_out["owner_email"] = employee.email

        result.append(eq_out)

    return result


@router.get(
    "/my",
    response_model=List[EquipmentOut],
    dependencies=[Depends(require_it_roles(["admin", "it_specialist", "employee"]))],
)
def list_my_equipment(
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
) -> List[Equipment]:
    return (
        db.query(Equipment)
        .filter(
            Equipment.current_owner_id == user.id, Equipment.status != "written_off"
        )
        .order_by(Equipment.name)
        .all()
    )


@router.post(
    "/scan-computer",
    response_model=EquipmentOut,
    dependencies=[Depends(require_it_roles(["admin", "it_specialist"]))],
)
def scan_computer_and_sync(
    payload: ScanComputerRequest,
    db: Session = Depends(get_db),
) -> EquipmentOut:
    """
    Сканировать ПК по имени или IP через WinRM-шлюз (учётка AD из интеграции),
    затем обновить соответствующую запись оборудования в Elements.
    """
    computer_name_or_ip = (payload.computer_name_or_ip or "").strip()
    if not computer_name_or_ip:
        raise HTTPException(status_code=400, detail="Укажите имя или IP компьютера")

    config = get_scan_config(db)
    gateway_host = (config.get("gateway_host") or "").strip()
    if not gateway_host:
        raise HTTPException(
            status_code=400,
            detail="Не настроен шлюз сканирования. Укажите scan_gateway_host в настройках (Интеграция AD).",
        )
    username = (config.get("username") or "").strip()
    password = config.get("password") or ""
    if not username or not password:
        raise HTTPException(
            status_code=400,
            detail="Не задана учётная запись AD для сканирования. Заполните ldap_bind_dn и ldap_bind_password в настройках.",
        )

    try:
        gateway_port = config.get("gateway_port")
        try:
            gateway_port = int(gateway_port) if gateway_port is not None else 5985
        except (TypeError, ValueError):
            gateway_port = 5985

        scan_result = run_scan(
            computer_name_or_ip=computer_name_or_ip,
            gateway_host=gateway_host,
            gateway_port=gateway_port,
            gateway_use_ssl=bool(config.get("gateway_use_ssl")),
            username=username,
            password=password,
        )
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except RuntimeError as e:
        # Логируем исходную ошибку (WinRM/сеть), чтобы в docker logs была видна причина
        cause = e.__cause__
        if cause:
            logger.warning(
                "Сканирование ПК: %s (исходная ошибка: %s)",
                e,
                cause,
                exc_info=False,
            )
        raise HTTPException(status_code=502, detail=str(e))
    except Exception as e:
        logger.exception("Ошибка при сканировании ПК: %s", e)
        raise HTTPException(
            status_code=500,
            detail=f"Ошибка сканирования: {e!s}. Проверьте логи сервера.",
        )

    try:
        sync_payload = EquipmentSyncFromScan(**scan_result)
        return sync_equipment_from_scan(sync_payload, db)
    except HTTPException:
        raise
    except Exception as e:
        logger.exception("Ошибка при обновлении оборудования после сканирования: %s", e)
        raise HTTPException(
            status_code=500,
            detail=f"Сканирование прошло, но не удалось обновить запись: {e!s}",
        )


@router.post(
    "/sync-from-scan",
    response_model=EquipmentOut,
    dependencies=[Depends(require_it_roles(["admin", "it_specialist"]))],
)
def sync_equipment_from_scan(
    payload: EquipmentSyncFromScan,
    db: Session = Depends(get_db),
) -> EquipmentOut:
    """
    Обновить оборудование данными от сканера ПК (по имени компьютера или по IP).
    Ищет запись по полю hostname или ip_address, обновляет характеристики и IP.
    """
    computer_name = (payload.computer_name or "").strip()
    ip_address = (payload.ip_address or "").strip() or None

    q = db.query(Equipment).filter(Equipment.category.in_(["computer", "server", "other"]))
    # Поиск: по hostname или по ip_address (если передан)
    if computer_name and ip_address:
        eq = q.filter(
            (Equipment.hostname == computer_name) | (Equipment.ip_address == ip_address)
        ).first()
    elif computer_name:
        eq = q.filter(Equipment.hostname == computer_name).first()
    elif ip_address:
        eq = q.filter(Equipment.ip_address == ip_address).first()
    else:
        raise HTTPException(status_code=400, detail="Укажите computer_name или ip_address")

    if not eq:
        raise HTTPException(
            status_code=404,
            detail=f"Оборудование с hostname '{computer_name}' или IP '{ip_address}' не найдено. "
            "Добавьте ПК в Elements и укажите имя компьютера (hostname) или IP.",
        )

    # Обновляем поля
    if computer_name:
        eq.hostname = computer_name
    if ip_address:
        eq.ip_address = ip_address
    if payload.serial_number is not None:
        eq.serial_number = payload.serial_number
    if payload.manufacturer is not None:
        eq.manufacturer = payload.manufacturer
    if payload.model is not None:
        eq.model = payload.model

    specs = dict(eq.specifications or {})
    if payload.cpu is not None:
        specs["cpu"] = payload.cpu
    if payload.ram is not None:
        specs["ram"] = payload.ram
    if payload.storage is not None:
        specs["storage"] = payload.storage
    if payload.os is not None:
        specs["os"] = payload.os
    if payload.disks is not None:
        specs["disks"] = payload.disks
    eq.specifications = specs if specs else None

    try:
        db.commit()
        db.refresh(eq)
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))

    result = EquipmentOut.model_validate(eq)
    if eq.room_id:
        room = db.query(Room).filter(Room.id == eq.room_id).first()
        if room:
            result.room_name = room.name
            if room.building_id:
                b = db.query(Building).filter(Building.id == room.building_id).first()
                if b:
                    result.building_name = b.name
    if eq.current_owner_id:
        owner = db.query(Employee).filter(Employee.id == eq.current_owner_id).first()
        if owner:
            result.owner_name = owner.full_name
            result.owner_email = owner.email
    return result


@router.get(
    "/export",
    dependencies=[Depends(require_it_roles(["admin", "it_specialist", "auditor"]))],
)
def export_equipment_excel(
    db: Session = Depends(get_db),
    status: Optional[str] = Query(None),
    category: Optional[str] = Query(None),
    owner_id: Optional[int] = Query(None),
    room_id: Optional[UUID] = Query(None),
    search: Optional[str] = Query(None),
    # Список колонок для выгрузки (через запятую)
    columns: Optional[str] = Query(None),
) -> StreamingResponse:
    """Экспорт каталога оборудования в Excel с выбором колонок."""
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl не установлен")

    # Все доступные колонки
    ALL_COLUMNS = {
        "inventory_number": "Инв. номер",
        "name": "Наименование",
        "category": "Категория",
        "status": "Статус",
        "manufacturer": "Производитель",
        "model": "Модель",
        "serial_number": "Серийный номер",
        "owner_name": "Ответственный",
        "owner_email": "Email ответственного",
        "department": "Отдел",
        "building_name": "Здание",
        "room_name": "Кабинет",
        "ip_address": "IP-адрес",
        "hostname": "Имя компьютера",
        "purchase_date": "Дата покупки",
        "warranty_until": "Гарантия до",
        "cost": "Стоимость",
    }

    CATEGORY_LABELS = {
        "computer": "Компьютер",
        "monitor": "Монитор",
        "printer": "Принтер",
        "network": "Сетевое оборудование",
        "server": "Сервер",
        "mobile": "Мобильное устройство",
        "peripheral": "Периферия",
        "other": "Прочее",
    }

    STATUS_LABELS = {
        "in_stock": "На складе",
        "in_use": "В использовании",
        "in_repair": "В ремонте",
        "written_off": "Списано",
    }

    # Определяем нужные колонки
    if columns:
        selected = [c.strip() for c in columns.split(",") if c.strip() in ALL_COLUMNS]
    else:
        selected = list(ALL_COLUMNS.keys())

    if not selected:
        selected = list(ALL_COLUMNS.keys())

    # Запрос оборудования (без пагинации — выгружаем всё)
    from backend.modules.hr.models.department import Department

    q = db.query(Equipment)
    if status:
        q = q.filter(Equipment.status == status)
    if category:
        q = q.filter(Equipment.category == category)
    if owner_id:
        q = q.filter(Equipment.current_owner_id == owner_id)
    if room_id:
        q = q.filter(Equipment.room_id == room_id)
    if search and search.strip():
        s = f"%{search.strip()}%"
        q = q.filter(
            or_(
                Equipment.name.ilike(s),
                Equipment.inventory_number.ilike(s),
                Equipment.serial_number.ilike(s),
            )
        )
    q = q.order_by(Equipment.category, Equipment.name)
    equipment_list = q.all()

    # Предзагружаем связанные данные
    room_ids = [eq.room_id for eq in equipment_list if eq.room_id]
    rooms_map: dict = {}
    if room_ids:
        rooms = db.query(Room).filter(Room.id.in_(room_ids)).all()
        building_ids = [r.building_id for r in rooms if r.building_id]
        buildings_map: dict = {}
        if building_ids:
            buildings = db.query(Building).filter(Building.id.in_(building_ids)).all()
            buildings_map = {b.id: b.name for b in buildings}
        rooms_map = {r.id: (r.name, buildings_map.get(r.building_id)) for r in rooms}

    owner_ids = [eq.current_owner_id for eq in equipment_list if eq.current_owner_id]
    owners_map: dict = {}
    departments_map: dict = {}
    if owner_ids:
        from backend.modules.hr.models.employee import Employee
        owners = (
            db.query(Employee)
            .filter(Employee.id.in_(owner_ids))
            .all()
        )
        dept_ids = [o.department_id for o in owners if o.department_id]
        if dept_ids:
            depts = db.query(Department).filter(Department.id.in_(dept_ids)).all()
            departments_map = {d.id: d.name for d in depts}
        owners_map = {
            o.id: (o.full_name, o.email, departments_map.get(o.department_id, ""))
            for o in owners
        }

    # Создаём книгу Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Оборудование"

    # Стиль заголовка
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    cell_font = Font(name="Calibri", size=10)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # Заголовки
    for col_idx, col_key in enumerate(selected, start=1):
        cell = ws.cell(row=1, column=col_idx, value=ALL_COLUMNS[col_key])
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align

    ws.row_dimensions[1].height = 30

    # Данные
    for row_idx, eq in enumerate(equipment_list, start=2):
        room_name, building_name, owner_name, owner_email, department = (
            None, None, None, None, None
        )
        if eq.room_id and eq.room_id in rooms_map:
            room_name, building_name = rooms_map[eq.room_id]
        if eq.current_owner_id and eq.current_owner_id in owners_map:
            owner_name, owner_email, department = owners_map[eq.current_owner_id]

        row_data: dict = {
            "inventory_number": eq.inventory_number or "",
            "name": eq.name or "",
            "category": CATEGORY_LABELS.get(eq.category, eq.category or ""),
            "status": STATUS_LABELS.get(eq.status, eq.status or ""),
            "manufacturer": eq.manufacturer or "",
            "model": eq.model or "",
            "serial_number": eq.serial_number or "",
            "owner_name": owner_name or "",
            "owner_email": owner_email or "",
            "department": department or "",
            "building_name": building_name or "",
            "room_name": room_name or "",
            "ip_address": eq.ip_address or "",
            "hostname": eq.hostname or "",
            "purchase_date": eq.purchase_date.strftime("%d.%m.%Y") if eq.purchase_date else "",
            "warranty_until": eq.warranty_until.strftime("%d.%m.%Y") if eq.warranty_until else "",
            "cost": float(eq.cost) if eq.cost is not None else "",
        }

        for col_idx, col_key in enumerate(selected, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row_data.get(col_key, ""))
            cell.font = cell_font
            cell.alignment = left_align

        # Чередование строк
        if row_idx % 2 == 0:
            alt_fill = PatternFill(start_color="EBF3FB", end_color="EBF3FB", fill_type="solid")
            for col_idx in range(1, len(selected) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = alt_fill

    # Авто-ширина колонок
    for col_idx, col_key in enumerate(selected, start=1):
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        header_len = len(ALL_COLUMNS[col_key])
        max_len = header_len
        for row_idx in range(2, len(equipment_list) + 2):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 50)

    # Заморозить первую строку
    ws.freeze_panes = "A2"

    # Добавляем автофильтр
    if equipment_list:
        ws.auto_filter.ref = ws.dimensions

    # Сохраняем в буфер
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"equipment_{timestamp}.xlsx"

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get(
    "/{equipment_id}",
    response_model=EquipmentOut,
    dependencies=[Depends(require_it_roles(["admin", "it_specialist", "employee", "auditor"]))],
)
def get_equipment(
    equipment_id: UUID,
    db: Session = Depends(get_db),
) -> EquipmentOut:
    eq = db.query(Equipment).filter(Equipment.id == equipment_id).first()
    if not eq:
        raise HTTPException(status_code=404, detail="Оборудование не найдено")
    
    # Формируем базовый ответ
    result = EquipmentOut.model_validate(eq)
    
    # Получаем информацию о владельце (сотруднике)
    if eq.current_owner_id:
        owner = db.query(Employee).filter(Employee.id == eq.current_owner_id).first()
        if owner:
            result.owner_name = owner.full_name
            result.owner_email = owner.email
    
    # Получаем информацию о кабинете и здании
    if eq.room_id:
        room = db.query(Room).filter(Room.id == eq.room_id).first()
        if room:
            result.room_name = room.name
            building = db.query(Building).filter(Building.id == room.building_id).first()
            if building:
                result.building_name = building.name
    
    # Получаем информацию о модели оборудования
    if eq.model_id:
        model = db.query(EquipmentModel).filter(EquipmentModel.id == eq.model_id).first()
        if model:
            result.model_name = model.name
            eq_type = db.query(EquipmentType).filter(EquipmentType.id == model.equipment_type_id).first()
            if eq_type:
                result.type_name = eq_type.name
                brand = db.query(Brand).filter(Brand.id == eq_type.brand_id).first()
                if brand:
                    result.brand_name = brand.name
    
    return result


@router.post(
    "/",
    response_model=EquipmentOut,
    status_code=201,
    dependencies=[Depends(require_it_roles(["admin", "it_specialist"]))],
)
def create_equipment(
    payload: EquipmentCreate,
    db: Session = Depends(get_db),
) -> Equipment:
    data = payload.model_dump()
    eq = Equipment(**data)
    db.add(eq)
    try:
        db.commit()
        db.refresh(eq)
    except Exception as e:
        db.rollback()
        err = str(e).lower()
        if "unique" in err and "inventory_number" in err:
            raise HTTPException(
                status_code=400,
                detail="Оборудование с таким инвентарным номером уже существует",
            )
        if "foreign key" in err or "23503" in err:
            raise HTTPException(
                status_code=400, detail="Некорректный владелец оборудования"
            )
        raise
    return eq


@router.patch(
    "/{equipment_id}",
    response_model=EquipmentOut,
    dependencies=[Depends(require_it_roles(["admin", "it_specialist"]))],
)
def update_equipment(
    equipment_id: UUID,
    payload: EquipmentUpdate,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
) -> Equipment:
    eq = db.query(Equipment).filter(Equipment.id == equipment_id).first()
    if not eq:
        raise HTTPException(status_code=404, detail="Оборудование не найдено")

    # Сохраняем старые значения для истории
    old_owner_id = eq.current_owner_id
    old_location_department = eq.location_department
    old_location_room = eq.location_room

    # Формируем старое местоположение
    old_location = None
    if old_location_department:
        if old_location_room:
            old_location = f"{old_location_department} - {old_location_room}"
        else:
            old_location = old_location_department

    # Обновляем поля
    update_data = payload.model_dump(exclude_unset=True)
    new_owner_id = update_data.get("current_owner_id")
    new_location_department = update_data.get("location_department")
    new_location_room = update_data.get("location_room")

    # Формируем новое местоположение
    new_location = None
    if new_location_department:
        if new_location_room:
            new_location = f"{new_location_department} - {new_location_room}"
        else:
            new_location = new_location_department

    # Применяем изменения
    for k, v in update_data.items():
        setattr(eq, k, v)

    try:
        db.commit()

        # Создаем запись в истории, если изменился владелец или местоположение
        if (new_owner_id != old_owner_id) or (new_location != old_location):
            history = EquipmentHistory(
                equipment_id=equipment_id,
                from_user_id=old_owner_id,
                to_user_id=new_owner_id,
                from_location=old_location,
                to_location=new_location,
                reason=None,  # Можно добавить в EquipmentUpdate если нужно
                changed_by_id=user.id,
            )
            db.add(history)
            db.commit()

        db.refresh(eq)
    except Exception as e:
        db.rollback()
        err = str(e).lower()
        if "unique" in err and "inventory_number" in err:
            raise HTTPException(
                status_code=400,
                detail="Оборудование с таким инвентарным номером уже существует",
            )
        if "foreign key" in err or "23503" in err:
            raise HTTPException(
                status_code=400, detail="Некорректный владелец оборудования"
            )
        raise
    return eq


@router.post(
    "/{equipment_id}/change-owner",
    response_model=EquipmentOut,
    dependencies=[Depends(require_it_roles(["admin", "it_specialist"]))],
)
def change_equipment_owner(
    equipment_id: UUID,
    payload: ChangeOwnerRequest,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
) -> Equipment:
    """Изменить владельца оборудования с созданием записи в истории"""
    eq = db.query(Equipment).filter(Equipment.id == equipment_id).first()
    if not eq:
        raise HTTPException(status_code=404, detail="Оборудование не найдено")

    # Сохраняем старые значения
    old_owner_id = eq.current_owner_id
    old_location_department = eq.location_department
    old_location_room = eq.location_room

    # Формируем старое местоположение
    old_location = None
    if old_location_department:
        if old_location_room:
            old_location = f"{old_location_department} - {old_location_room}"
        else:
            old_location = old_location_department

    # Обновляем оборудование
    eq.current_owner_id = payload.new_owner_id
    if payload.new_location_department is not None:
        eq.location_department = payload.new_location_department
    if payload.new_location_room is not None:
        eq.location_room = payload.new_location_room

    # Формируем новое местоположение
    new_location = None
    if eq.location_department:
        if eq.location_room:
            new_location = f"{eq.location_department} - {eq.location_room}"
        else:
            new_location = eq.location_department

    try:
        db.commit()

        # Создаем запись в истории
        history = EquipmentHistory(
            equipment_id=equipment_id,
            from_user_id=old_owner_id,
            to_user_id=payload.new_owner_id,
            from_location=old_location,
            to_location=new_location,
            reason=payload.reason,
            changed_by_id=user.id,
        )
        db.add(history)
        db.commit()
        db.refresh(eq)
    except Exception as e:
        db.rollback()
        err = str(e).lower()
        if "foreign key" in err or "23503" in err:
            raise HTTPException(
                status_code=400, detail="Некорректный владелец оборудования"
            )
        raise
    return eq


@router.delete(
    "/{equipment_id}",
    status_code=200,
    dependencies=[Depends(require_it_roles(["admin", "it_specialist"]))],
)
def delete_equipment(
    equipment_id: UUID,
    db: Session = Depends(get_db),
) -> dict:
    eq = db.query(Equipment).filter(Equipment.id == equipment_id).first()
    if not eq:
        raise HTTPException(status_code=404, detail="Оборудование не найдено")
    db.delete(eq)
    db.commit()
    return {"message": "Оборудование удалено"}


@router.get(
    "/{equipment_id}/consumables",
    dependencies=[Depends(require_it_roles(["admin", "it_specialist", "employee"]))],
)
def get_equipment_consumables(
    equipment_id: UUID,
    db: Session = Depends(get_db),
) -> List[dict]:
    """Получить расходные материалы для оборудования (через модель оборудования)."""
    eq = db.query(Equipment).filter(Equipment.id == equipment_id).first()
    if not eq:
        raise HTTPException(status_code=404, detail="Оборудование не найдено")

    if not eq.model_id:
        return []

    # Получаем расходники модели с информацией о наличии
    model_consumables = (
        db.query(ModelConsumable, Consumable)
        .outerjoin(Consumable, ModelConsumable.consumable_id == Consumable.id)
        .filter(
            ModelConsumable.model_id == eq.model_id,
            ModelConsumable.is_active == True,
        )
        .order_by(ModelConsumable.name)
        .all()
    )

    result = []
    for mc, consumable in model_consumables:
        item = {
            "consumable_id": str(mc.consumable_id) if mc.consumable_id else str(mc.id),
            "consumable_name": mc.name,
            "consumable_model": mc.part_number,
            "consumable_type": mc.consumable_type,
            "quantity_in_stock": consumable.quantity_in_stock if consumable else 0,
            "min_quantity": consumable.min_quantity if consumable else 0,
            "is_low_stock": (consumable.quantity_in_stock <= consumable.min_quantity)
            if consumable
            else False,
        }
        result.append(item)

    return result


@router.get(
    "/{equipment_id}/licenses/",
    dependencies=[Depends(require_it_roles(["admin", "it_specialist", "employee"]))],
)
def get_equipment_licenses(
    equipment_id: UUID,
    db: Session = Depends(get_db),
) -> List[dict]:
    """Получить лицензии, привязанные к оборудованию."""
    eq = db.query(Equipment).filter(Equipment.id == equipment_id).first()
    if not eq:
        raise HTTPException(status_code=404, detail="Оборудование не найдено")

    # Получаем активные привязки лицензий к этому оборудованию
    assignments = (
        db.query(LicenseAssignment, SoftwareLicense)
        .join(SoftwareLicense, LicenseAssignment.license_id == SoftwareLicense.id)
        .filter(
            LicenseAssignment.equipment_id == equipment_id,
            LicenseAssignment.released_at.is_(None),  # Только активные
        )
        .order_by(LicenseAssignment.assigned_at.desc())
        .all()
    )

    result = []
    for assignment, license in assignments:
        result.append({
            "id": str(license.id),
            "software_name": license.software_name,
            "vendor": license.vendor,
            "license_type": license.license_type,
            "expires_at": license.expires_at.isoformat() if license.expires_at else None,
            "assigned_at": assignment.assigned_at.isoformat() if assignment.assigned_at else None,
        })

    return result
